# -*- coding: utf-8 -*-
"""Fill column E 动作库描述（全） from nameEn and sheet metadata."""
import re
import shutil
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

XLSX = Path(r"G:\AI_Workspaces\动作库\动作库清单.xlsx")
OUT = XLSX

# 仅写「从【标准动作描述】起」的定制正文；抬头四行由程序按表内字段拼装。
BODY_OVERRIDES = {
    "Band Calf Raise": """

【标准动作描述】
起始位：
站姿，双脚与肩同宽，前脚掌踩住弹力带中段，双手握住两端并拉起至身体两侧（或肩侧）形成持续阻力。躯干立直，核心收紧，膝关节微屈不锁死。

离心阶段（下放）：
吸气，脚跟缓慢下落至接近地面，保持弹力带不断张力，避免突然卸力。

向心阶段（发力）：
呼气，前脚掌蹬地，脚跟向上抬至最高可控位置，感受小腿后侧收缩，身体不前后晃动。

顶峰阶段：
顶端停 0.5-1 秒，主动夹紧小腿，再进入下一次。

【呼吸与节奏】
呼吸： 上提呼气，下放吸气。
节奏： 2-1-2-0（上提2秒-顶端停1秒-下放2秒-底端不停顿）。

【关键技术要点】
全程重心稳在前脚掌与脚跟连线上，不内扣外翻。
动作以踝关节跖屈为主，不借屈膝弹动。
下放和上提都要可控，避免“抖动式”半程。
阻力以能完成全程幅度为准，先质量后加阻力。

【常见错误与纠正】
错误： 身体上下弹，靠惯性完成。
纠正口令： “慢慢提、慢慢放，像电梯别像弹簧。”

错误： 幅度太小，只做半程。
纠正口令： “脚跟先放低，再踮到最高。”

错误： 脚踝内扣/外翻，发力跑偏。
纠正口令： “脚掌三点着力，膝盖对准脚尖。”

【负荷与做组建议】
新手技术/耐力： 3-4组 × 15-20次，RIR 2-3。
进阶增肌： 4-5组 × 12-18次，RIR 1-2。
力量耐力倾向： 4组 × 10-15次（更高阻力），RIR 1-2。
组间休息： 45-90秒。
""".strip(),
}

from body_overrides_extra import BODY_OVERRIDES_EXTRA

BODY_OVERRIDES.update(BODY_OVERRIDES_EXTRA)

STARS = {
    "Beginner": "★★☆☆☆（2/5）",
    "Novice": "★★☆☆☆（2/5）",
    "Intermediate": "★★★☆☆（3/5）",
    "Advanced": "★★★★☆（4/5）",
}


def norm_en(name: str) -> str:
    if not isinstance(name, str):
        return ""
    s = name.strip()
    s = re.sub(r"_black_sticker$", "", s, flags=re.I)
    return s


def cat_line(typ: str) -> str:
    t = (typ or "").strip()
    if t in ("Stretches",):
        return "分类： 拉伸（主）；热身（可选）；有氧（否）；无氧（否）"
    if t in ("Cardio",):
        return "分类： 有氧（主）；热身（可选）；无氧（否）；拉伸（否）"
    if t in ("Recovery", "Yoga"):
        return "分类： 热身（主，恢复/柔韧）；拉伸（可选）；有氧（否）；无氧（否）"
    # default strength / accessories
    return "分类： 无氧（主）；热身（可选）；有氧（否）；拉伸（否）"


def minor_default(main: str) -> str:
    _ = (main or "").strip()
    return "次要肌群： 邻近关节稳定肌群、核心稳定肌群"


def difficulty_star(d: str) -> str:
    d = (d or "").strip()
    return STARS.get(d, "★★★☆☆（3/5）")


def safe_str(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    return str(x).strip()


def build_description(row) -> str:
    en = norm_en(safe_str(row.get("nameEn")))
    cn = safe_str(row.get("nameEn_cn")) or en
    typ = safe_str(row.get("type"))
    diff_en = safe_str(row.get("difficulty"))
    main_m = safe_str(row.get("muscleGroup_cn")) or "目标肌群"

    star = difficulty_star(diff_en)

    head = f"{cn}（{en}）\n{cat_line(typ)}\n难度： {star}\n主要肌群： {main_m}\n{minor_default(main_m)}"

    if en in BODY_OVERRIDES:
        return f"{head}\n\n{BODY_OVERRIDES[en]}"

    if typ == "Stretches":
        body = f"""

【标准动作描述】
起始位：
按图示进入拉伸起始姿势，脊柱尽量保持中立，支撑侧关节稳定、不耸肩。

离心阶段（加深）：
呼气或自然呼吸，在无痛前提下缓慢加深拉伸幅度，幅度以肌肉有牵拉感且无刺痛为准。

向心阶段（还原）：
吸气，缓慢回到起始位或略微放松，避免弹振式猛拉。

顶峰阶段：
在最大舒适角度停留，感受目标肌群牵拉；保持均匀呼吸，不憋气。

【呼吸与节奏】
呼吸： 缓慢加深时呼气更顺畅者可稍呼气；以能全程放松肩颈为准，不强制憋气。
节奏： 30-60秒静态保持/侧（可重复2-3轮）；或 4-0-4-0（加深4秒-停留-还原4秒）。

【关键技术要点】
以目标肌群牵拉为主，避免用邻近关节代偿“假幅度”。
肩、颈、腰如有不适，先减小幅度或停止。
身体放松但核心轻收，保护腰椎。

【常见错误与纠正】
错误： 弹振式拉扯、追求撕裂感。
纠正口令： “慢一点，像橡皮筋慢慢拉长，别弹。”

错误： 憋气、耸肩、躯干扭成别的方向。
纠正口令： “肩膀下沉，呼吸续上，幅度先减半。”

【负荷与做组建议】
新手技术/耐力： 每侧2-3组 × 静态20-30秒，强度以无痛牵拉为准。
进阶： 每侧2-4组 × 静态30-45秒。
力量耐力倾向： 不适用；以主观舒适度与关节安全为准。
组间休息： 20-45秒。
""".strip()
        return f"{head}\n\n{body}"

    if typ == "Cardio":
        body = f"""

【标准动作描述】
起始位：
按器械/课程要求站稳或就位，躯干中立，核心微收，目视前方。

离心阶段（还原/回摆）：
有控制地回到起始姿势，膝关节对齐脚尖方向，避免关节撞击感。

向心阶段（发力/推蹬）：
节奏稳定地完成蹬伸或摆臂，逐步提高心率，避免只靠惯性猛冲。

顶峰阶段：
在动作末端保持1拍节奏感与稳定，再进入下一次循环。

【呼吸与节奏】
呼吸： 以节奏化鼻吸口呼为主，不要长时间憋气。
节奏： 有氧区间以可持续对话或短句为宜；若编排课程按口令执行。

【关键技术要点】
足底三点支撑稳定，膝关节轨迹与脚尖一致。
循序渐进升强度，先动作质量再速度。
补水与热身充分，出现眩晕、胸闷立即停止。

【常见错误与纠正】
错误： 身体松散、膝盖内扣或过度前移。
纠正口令： “核心收一点，膝盖对准脚尖，别砸地。”

错误： 为了追求速度失去控制。
纠正口令： “先稳后快，步子别飘。”

【负荷与做组建议】
新手技术/耐力： 10-20分钟中等强度，自我感觉轻松-中等。
进阶： 20-40分钟；可穿插高强度间歇（遵守个人健康状况）。
力量耐力倾向： 不适用或按课程间歇执行。
组间休息： 间歇训练按课表；持续有氧以补水为主。
""".strip()
        return f"{head}\n\n{body}"

    # Generic strength / machine / free weight / bodyweight
    body = f"""

【标准动作描述】
起始位：
按动作模式就位（站姿/坐姿/仰卧/俯身等），肩胛与核心预先稳定，关节排列中立；抓握牢固，目视前方或按动作要求注视固定点。

离心阶段（下放）：
吸气或按动作节奏换气，在目标肌群可控的前提下缓慢下放/展开，全程保持张力，避免突然卸力或弹震。

向心阶段（发力）：
呼气为主完成主要发力阶段，想象“目标肌群主导、相邻关节协同”，轨迹稳定一致，不借大幅度甩动代偿。

顶峰阶段：
在动作顶端短暂停留0.5-1秒，主动收紧目标肌群，确认肩关节/腰椎无挤压不适后进入下一次。

【呼吸与节奏】
呼吸： 发力阶段呼气更常见；全程避免过度憋气导致头晕。
节奏： 2-1-2-0（向心2秒—顶端停1秒—离心2秒—底端不停顿）；大重量力量训练可采用2-1-1-0或按专项节奏。

【关键技术要点】
肩胛、腰腹、髋膝踝排列优先，再追求负重与次数。
动作质量优于盲目冲重量；幅度以无痛、可重复为标准。
组间补呼吸，避免连续组间过度憋气。

【常见错误与纠正】
错误： 代偿借力（甩动身体/弹跳/耸肩）导致轨迹散。
纠正口令： “先稳住躯干，再完成动作，幅度可以减半。”

错误： 关节排列丢失（塌腰、含胸、膝盖内扣等）。
纠正口令： “肋骨收回、核心绷紧，膝盖跟脚尖一条线。”

【负荷与做组建议】
新手技术/耐力： 3-4组 × 8-15次，RIR 2-3。
进阶增肌： 4-5组 × 6-12次，RIR 1-2。
力量耐力倾向： 4-6组 × 3-6次（大重量需保护与安全杠），RIR 1-2。
组间休息： 增肌60-120秒；力量120-180秒；代谢类30-60秒。
""".strip()
    return f"{head}\n\n{body}"


def main():
    df = pd.read_excel(XLSX)
    texts = [build_description(df.iloc[i].to_dict()) for i in range(len(df))]

    backup = XLSX.with_name(XLSX.stem + "_backup_" + datetime.now().strftime("%Y%m%d_%H%M%S") + XLSX.suffix)
    shutil.copy2(XLSX, backup)
    print("Backup:", backup)

    wb = load_workbook(XLSX)
    ws = wb.active
    # Column E = 5 is 动作库描述（全）
    header_e = ws.cell(row=1, column=5).value
    if header_e is None or str(header_e).strip() == "":
        ws.cell(row=1, column=5, value="动作库描述（全）")

    wrap = Alignment(wrap_text=True, vertical="top")
    for r, txt in enumerate(texts, start=2):
        c = ws.cell(row=r, column=5, value=txt)
        c.alignment = wrap
    ws.column_dimensions["E"].width = 60
    if ws.cell(row=1, column=5).alignment is None:
        ws.cell(row=1, column=5).alignment = Alignment(wrap_text=True, vertical="center")

    wb.save(OUT)
    print("Wrote", len(texts), "rows to column E ->", OUT)


if __name__ == "__main__":
    main()
